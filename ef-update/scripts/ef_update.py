"""
EF Master Update Script
Updates Task sheet, Todays Work sheet, Dashboard, and resets the batch file.
Usage: python ef_update.py <batch_file> [ef_master_path]
"""

import sys
import re
import os
from datetime import datetime
from copy import copy

try:
    from openpyxl import load_workbook
    import pandas as pd
except ImportError:
    print("ERROR: Required packages missing. Run: pip install openpyxl pandas")
    sys.exit(1)


def extract_share_link(ops_done):
    """Extract R6 Share Link URL from Operations Done text."""
    if pd.isna(ops_done) or not ops_done:
        return None
    match = re.search(r'R6: Share Link: (https://maps\.app\.goo\.gl/\S+)', str(ops_done))
    return match.group(1).strip() if match else None


def main():
    if len(sys.argv) < 2:
        print("Usage: python ef_update.py <batch_file> [ef_master_path]")
        print("Example: python ef_update.py C:\\Users\\nisha\\OneDrive\\Desktop\\80.xlsx")
        sys.exit(1)

    batch_path = sys.argv[1]
    ef_master_path = sys.argv[2] if len(sys.argv) > 2 else r"C:\Users\nisha\OneDrive\Desktop\EF New\EF Master.xlsx"

    if not os.path.exists(batch_path):
        print(f"ERROR: Batch file not found: {batch_path}")
        sys.exit(1)
    if not os.path.exists(ef_master_path):
        print(f"ERROR: EF Master not found: {ef_master_path}")
        sys.exit(1)

    batch_name = os.path.basename(batch_path)
    today_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_str = datetime.now().strftime("%d/%m/%Y")
    now_str = datetime.now().strftime("%H:%M:%S")

    # ========== READ BATCH FILE ==========
    print(f"Reading batch file: {batch_name}")
    df_batch = pd.read_excel(batch_path)
    total_entries = len(df_batch)

    # Extract share links from Operations Done BEFORE clearing
    share_links = {}
    share_count = 0
    for i, row in df_batch.iterrows():
        email = row.get('Email', '')
        ops_done = row.get('Operations Done', '')
        link = extract_share_link(ops_done)
        if email and link:
            share_links[email] = link
            share_count += 1

    print(f"Entries: {total_entries}, Share links recovered: {share_count}")

    # ========== UPDATE EF MASTER ==========
    print(f"\nUpdating EF Master: {os.path.basename(ef_master_path)}")
    wb = load_workbook(ef_master_path)

    # ---------- 1. TASK SHEET ----------
    task_ws = wb['Task']

    # Find last data row
    last_num = 0
    last_row = 0
    for r in range(task_ws.max_row, 1, -1):
        val = task_ws.cell(row=r, column=1).value
        if val is not None and str(val).isdigit():
            last_row = r
            last_num = int(val)
            break

    # Copy formatting from last row
    ref_styles = []
    for col in range(1, 9):
        c = task_ws.cell(row=last_row, column=col)
        ref_styles.append({
            'font': copy(c.font) if c.font else None,
            'alignment': copy(c.alignment) if c.alignment else None,
            'border': copy(c.border) if c.border else None,
            'fill': copy(c.fill) if c.fill else None,
            'number_format': c.number_format,
        })

    # Get date format from existing entries
    date_format = ref_styles[6]['number_format'] if ref_styles[6]['number_format'] != 'General' else 'dd/mm/yyyy'

    # Add entries
    start_num = last_num + 1
    for i, row_data in df_batch.iterrows():
        new_num = last_num + 1 + i
        new_row = last_row + 1 + i
        email = row_data.get('Email', '')
        note = share_links.get(email, None)

        values = [
            new_num,
            row_data.get('GMB Name', ''),
            row_data.get('GMB URL', ''),
            row_data.get('Review Text', ''),
            'Done',
            email,
            today_date,
            note,
        ]

        for col_idx, val in enumerate(values):
            cell = task_ws.cell(row=new_row, column=col_idx + 1, value=val)
            s = ref_styles[col_idx]
            if s['font']: cell.font = copy(s['font'])
            if s['alignment']: cell.alignment = copy(s['alignment'])
            if s['border']: cell.border = copy(s['border'])
            if s['fill']: cell.fill = copy(s['fill'])
            if col_idx == 6:  # Date column
                cell.number_format = date_format

    end_num = last_num + total_entries
    print(f"  Task: Added #{start_num} to #{end_num}")

    # Update merged headers
    try:
        task_ws.unmerge_cells('A1:H1')
    except (KeyError, ValueError):
        pass
    task_ws.cell(row=1, column=1).value = 'EF Auto Live - Daily Post Tracker'
    task_ws.merge_cells('A1:H1')

    try:
        task_ws.unmerge_cells('A2:H2')
    except (KeyError, ValueError):
        pass
    task_ws.cell(row=2, column=1).value = f'Last Updated: {today_str} {now_str}'
    task_ws.merge_cells('A2:H2')

    # ---------- 2. TODAYS WORK SHEET ----------
    tw_ws = wb['Todays Work']
    tw_ws.cell(row=1, column=1).value = f"Today's Work \u2014 {today_str}"

    # Build GMB URL lookup from batch
    gmb_lookup = {}
    for i, row_data in df_batch.iterrows():
        url = row_data.get('GMB URL', '')
        if url:
            gmb_lookup[url] = row_data.get('Email', '')

    # Mark matching entries as Done
    marked_done = 0
    for r in range(4, tw_ws.max_row + 1):
        gmb_link = tw_ws.cell(row=r, column=3).value
        if gmb_link in gmb_lookup:
            tw_ws.cell(row=r, column=6).value = 'Done'
            tw_ws.cell(row=r, column=7).value = gmb_lookup[gmb_link]
            tw_ws.cell(row=r, column=8).value = today_date
            marked_done += 1

    # Delete Done rows (bottom to top)
    rows_to_delete = []
    for r in range(tw_ws.max_row, 3, -1):
        if tw_ws.cell(row=r, column=6).value == 'Done':
            rows_to_delete.append(r)

    for r in rows_to_delete:
        tw_ws.delete_rows(r)

    # Renumber remaining
    remaining = 0
    for r in range(4, tw_ws.max_row + 1):
        if tw_ws.cell(row=r, column=2).value is not None:
            remaining += 1
            tw_ws.cell(row=r, column=1).value = remaining

    print(f"  Todays Work: {marked_done} marked Done, {len(rows_to_delete)} removed, {remaining} remaining")

    # ---------- 3. DASHBOARD ----------
    dash_ws = wb['Dashboard']
    gmb_urls = set(df_batch['GMB URL'].dropna().values)

    dash_updated = 0
    for r in range(2, dash_ws.max_row + 1):
        gmb_link = dash_ws.cell(row=r, column=4).value
        if gmb_link in gmb_urls:
            curr_rev = dash_ws.cell(row=r, column=9).value
            if curr_rev is not None:
                try:
                    dash_ws.cell(row=r, column=9).value = float(curr_rev) + 1
                    dash_updated += 1
                except (ValueError, TypeError):
                    pass

    print(f"  Dashboard: {dash_updated} Curr Reviews updated (+1)")

    # Save EF Master
    try:
        wb.save(ef_master_path)
        print(f"  EF Master saved!")
    except PermissionError:
        print(f"\n  ERROR: Cannot save - file is open in Excel. Close it and retry.")
        sys.exit(1)

    # ========== RESET BATCH FILE ==========
    print(f"\nResetting batch file: {batch_name}")
    wb_batch = load_workbook(batch_path)
    ws_batch = wb_batch.active

    # Find columns to clear by header name
    headers = {}
    for col in range(1, ws_batch.max_column + 1):
        val = ws_batch.cell(row=1, column=col).value
        if val:
            headers[val] = col

    clear_cols = [
        'Status', 'Error Message', 'Operations Done', 'Share Link', 'Date',
        'Op1: Change Password', 'Op2: Recovery Phone', 'Op3: Recovery Email',
        'Op4: Authenticator', 'Op5: Backup Codes', 'Op6: 2FA Phone',
        'Op7: Remove Devices', 'Op8: Change Name', 'Live Check Status',
    ]
    # Add unnamed columns
    clear_cols.extend([k for k in headers if k.startswith('Unnamed')])

    cleared = 0
    for col_name in clear_cols:
        if col_name in headers:
            col_idx = headers[col_name]
            for row in range(2, ws_batch.max_row + 1):
                ws_batch.cell(row=row, column=col_idx).value = None
            cleared += 1

    try:
        wb_batch.save(batch_path)
        print(f"  Cleared {cleared} output columns, input data intact")
    except PermissionError:
        print(f"  ERROR: Cannot save batch file - it may be open. Close and retry.")
        sys.exit(1)

    # ========== SUMMARY ==========
    print(f"\n{'='*40}")
    print(f"  EF Master Update Complete!")
    print(f"{'='*40}")
    print(f"  Batch: {batch_name} ({total_entries} entries)")
    print(f"  Share links: {share_count} recovered")
    print(f"  Task: #{start_num} to #{end_num} added")
    print(f"  Todays Work: {len(rows_to_delete)} removed, {remaining} pending")
    print(f"  Dashboard: {dash_updated} reviews updated")
    print(f"  Batch file: Reset for reuse")
    print(f"{'='*40}")


if __name__ == '__main__':
    main()
